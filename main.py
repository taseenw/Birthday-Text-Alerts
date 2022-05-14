# Author: Taseen Waseq
# Date Created: 2022-05-07
# This program will send the user a text message if it is their friend/familys birthday, drawing the information from a basic excel sheet consisting of names and birthdays

from twilio.rest import Client
from openpyxl import load_workbook
from datetime import date

#Function sending the text message using twilio library function, and twilio credentials
def text(msg):
    account_sid = "ENTER YOUR ACCOUNT SID HERE"
    auth_token = "ENTER YOUR AUTH TOKEN HERE"
    client = Client(account_sid, auth_token)

    client.messages \
                        .create(
                            body = msg,
                            from_ = '+YOUR TWILIO PHONE NUMBER',
                            to = '+YOUR VERIFIED PHONE NUMBER'
                        )

def main():
    #Data retrieval from Excel Sheet
    workbook = load_workbook("birthdayList.xlsx")
    currentWS = workbook[workbook.sheetnames[0]]
    nameCol = currentWS['A']
    birthdayCol = currentWS['B']

    #Number of entries: could use either column as both should be equal lengths
    entryQ = len(birthdayCol)

    #Populate the name and birthday list with the excel sheet values (Excel date format chr 0-9 is identical to pythond date format)
    nameList = []
    birthdayList = []
    for x in range(entryQ):
        curName = str(nameCol[x].value)
        curBirthday = str(birthdayCol[x].value)

        nameList.append(curName)
        birthdayList.append(curBirthday[0:10])

    #Set current date
    dStr = str(date.today())

    #If present date's string is present in list of birthday strings
    if dStr in birthdayList:
        birthdayName = nameList[birthdayList.index(dStr)]
        message ="It is "+birthdayName+"'s Birthday Today!"
    else:
        message = ""

    if message != "":
        text(message)

if __name__ == "__main__":
    main()